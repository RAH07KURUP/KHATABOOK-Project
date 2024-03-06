# To import libraries
from pandas import *
from openpyxl import load_workbook
import datetime as dt
from numpy import *
from ssl import *
from tkinter import *
from tkinter import messagebox
from smtplib import *
from PySimpleGUI import *
import sys
# To show the caution window
d = Tk()
d.geometry("150x100")
messagebox.showinfo('CAUTION!!','Keep both the User excel sheet and Khata excel sheet closed for the smooth functioning of the Program:)')
# To read the excel sheet as user and ks
user=read_excel(R"D:\android\downloads\Book2.xlsx")
ks=read_excel(R"D:\android\downloads\Book1.xlsx")

# To open a yes,no window
theme('LightGrey')
layout = [
    [Text("Are you new user?")],
    [Yes(), No()]
]
dialog = Window(" ", layout)
event,values = dialog.read()

if event==WIN_CLOSED:
    sys.exit()
    
# when 'yes' clicked.
if event == 'Yes' :

    def clean(x):
        wb=load_workbook(x)
        sh=wb['Sheet1']
        for i in range(2,sh.max_row+1):
            sh.delete_rows(i)
        wb.save(x)
        return None
            
    clean(R"D:\android\downloads\Book1.xlsx")
    clean(R"D:\android\downloads\Book2.xlsx")
    F=0 #Flag element
    # To enter the entries in excel sheet
    theme('Purple')
    layout = [
        [Text('Enter the personal Details:')],
        [Text('Username', size=(18, 1)), InputText(key='USERNAME')],
        [Text('valid Email-id', size=(18, 1)), InputText(key='EMAIL-ID')],
        [Text('E-mail Password', size=(18, 1)), InputText(key='PASSWORD')],
        [Text('valid Paytm No.(10 Digit)', size=(18, 1)), InputText(key='PAYTM NO')],
        [Text('Bank Account no.', size=(18, 1)), InputText(key='BANK ACCOUNT NO')],
        [Text('Max Debt Allowed(int)', size=(18, 1)), InputText(key='MAXIMUM PENDING AMOUNT')],
        [Text('Permissible Time(days)', size=(18, 1)), InputText(key='PERMISSIBLE TIME')],
        [Text('Organisation Name', size=(18, 1)), InputText(key='ORGANISATION')],
        [Button('Clear'), Submit(), Exit()]
    ]
    form = Window('PERSONAL DETAILS', layout)
#if 'No' is clicked     
if event == 'No' :
    F=1 # Flag element

    # To verify the user details
    theme('DarkBrown')
    layout = [
        [Text('Enter your login Details:')],
        [Text('Paytm-No', size=(15, 1)), InputText(key='Paytm')],
        [Text('E-mail Password', size=(15, 1)), InputText(key='Password')],
        [Button('Clear'), Submit(), Exit()]

    ]
    form = Window('VERIFICATION', layout)

# We define a clear function for clear all the entries of form   
def clear():
    for key in values:
        form[key]('')
    return None

while True:
    event, values = form.read()
    if event == "Clear":
        clear()
# If 'Submit' button is clicked
    if event == "Submit":
        if(F==0):
            wb=load_workbook(R"D:\android\downloads\Book2.xlsx")
            data=[values['USERNAME'],values['EMAIL-ID'],values['PASSWORD'],values['PAYTM NO'],values['BANK ACCOUNT NO'],values['MAXIMUM PENDING AMOUNT'],values['PERMISSIBLE TIME'],values['ORGANISATION']]
            data=[data[0],data[1],data[2],int(data[3]),data[4],int(data[5]),int(data[6]),data[7]]
            sheet=wb['Sheet1']
            sheet.append(data)
            wb.save(R"D:\android\downloads\Book2.xlsx")
            d.geometry('200x200')
            messagebox.showinfo('Congratulations '+values['USERNAME'],'You are successfuly registered!' )
            break
        else:
            d=[values['Paytm'],values['Password']]
            user=read_excel(R"D:\android\downloads\Book2.xlsx")
            if(int(d[0])!=user['PAYTM NO'][0] or d[1]!=user['PASSWORD'][0]):
                popup('INVALID CREDENTIALS!!','It seems that you have entered the wrong details. You can try again or may register as a new user.')
                form.close()
                dialog.close()
                sys.exit()
            else:
                messagebox.showinfo('','Welcome Back '+user['USERNAME'][0])
                break
                
    if event == "Exit" or event==WIN_CLOSED:
        form.close()
        dialog.close()
        sys.exit()

form.close()
dialog.close()

    
theme('LightGrey')
layout = [
    [Text('Do you want to make entries in Khata-Sheet?')],
    [Yes(), No()]
]
dialog = Window("Khata-Sheet Update", layout)
event, values = dialog.read()
if event == 'Yes':
    theme('Purple')
    layout = [
        [Text('Enter the Customer Details:')],
        [Text('Name', size=(18, 1)), InputText(key='NAME')],
        [Text('valid Email-Id', size=(18, 1)), InputText(key='EMAIL-ID')],
        [Text('Mob-no. (10 Digit)', size=(18, 1)), InputText(key='MOB NO')],
        [Text('Pending Amount (Integer)', size=(18, 1)), InputText(key='PA')],
        [Text('Debted on (dd-mm-yyyy)', size=(18, 1)), InputText(key='PS')],
        [Button('Clear'), Submit(), Exit()]
    ]

    form = Window('Customer Details ', layout)


    while True:
        event, values = form.read()
        if event == "Clear":
            clear()

        if event == "Submit":
            wb=load_workbook(R"D:\android\downloads\Book1.xlsx")
            data=[values['NAME'],values['EMAIL-ID'],values['MOB NO'],values['PA'],values['PS']]
            data=[data[0],data[1],int(data[2]),int(data[3]),dt.datetime.strptime((data[4]+' 00:00:00'),'%d-%m-%Y %H:%M:%S')]
            sheet=wb['Sheet1']
            sheet.append(data)
            wb.save(R"D:\android\downloads\Book1.xlsx")

            popup('Data Saved!')
            clear()
        if event == "Exit" or event==WIN_CLOSED:
            break

    form.close()
dialog.close()

#To read the Khata sheet
user=read_excel(R"D:\android\downloads\Book2.xlsx")
ks=read_excel(R"D:\android\downloads\Book1.xlsx")

#Funtion for sending emails
def email(cust_id,name,sub):
    #Storing email addresses of sender,receiver and also sender password
    sender=user['EMAIL-ID'][0]
    receiver=cust_id
    password=user['PASSWORD'][0]

    #BODY OF THE MESSAGE
    body='Respected '+str(ks['NAME'][i])+'\nYou have not paid an amount of '+str(ks['PA'][i])+' to '+str(user['ORGANISATION'][0])+' owned by '+user['USERNAME'][0]+'. Please pay asap as you have exceeded the Debt Limit.You can pay to-\n'+'PAYTM NO:\t'+str(user['PAYTM NO'][0])+'\nACCOUNT NO: '+str(user['BANK ACCOUNT NO'][0])
    #THE MESSAGE
    message=f"Subject:{sub}\nTo:{receiver}\n{body}"
    #Create a secure connection with the server and send the email
    #Create the secure socket layer(SSL) context
    SSL_context=create_default_context()
    #Create the secure SMTP connection
    server=SMTP_SSL(host='smtp.gmail.com',port=465,context=SSL_context)

    #Logging to the user mail id
    server.login(sender,password)

    #To send the e-mail
    server.sendmail(sender,receiver,message)

#Present Date
date=dt.datetime.now()

l,f={},0
#Khatasheet is checked for alert cases
for i in range(len(ks)):
    pa=ks['PA'][i]
    ps=(ks['PS'][i])
    PS=str(ps-date)
    d=''
    for j in range(1,len(PS)):
        if(PS[j]==' '):
            break
        d+=PS[j]
    PS=int(d)

    if(pa>user['MAX PENDING AMOUNT'][0] or PS>user['PERMISSIBLE TIME'][0]):
        email(ks['EMAIL-ID'][i],ks['NAME'][i],'PAYMENT PENDING')
        l[ks['NAME'][i]]=ks['EMAIL-ID'][i]
        f=1

if(f==1):
    d=Tk()
    d.geometry("200x200")
    s1 = 'The following customer(s) have exceeded the Debt limit:\n'+str(l)+"\nReminder Mail has been sent to each of them for payment"
    s2="Remember to Update the Khata Sheet after any PAYMENT is received"
    messagebox.showinfo("PENDING PAYMENT(S)",s1)
    messagebox.showinfo("NOTE:", s2)
else:
    d = Tk()
    d.geometry("150x100")
    messagebox.showinfo("CONGRATULATIONS!!!", "Your Khata is going fine. No need to worry:)")
    
